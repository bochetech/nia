"""
build_experiences_rag.py
========================
Convierte activities (1).xlsx en un documento Markdown estructurado para RAG,
orientado a asesoría: "¿qué experiencia es mejor para mí?".

Genera: data/seed/knowledge/experiences_advisor.md
"""
from __future__ import annotations

import re
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

# ── Configuración ──────────────────────────────────────────────

XLSX = Path(__file__).parent.parent / "activities (1).xlsx"
OUT  = Path(__file__).parent.parent / "data/seed/knowledge/experiences_advisor.md"

# Categorías a omitir (productos caja/add-on sin experiencia real)
SKIP_KEYWORDS = [
    "(Caja)", "Jugo de Uva", "Almuerzo Bodega 1883 (Caja)",
    "TTOO",  # tarifas operador turístico, no relevantes para el cliente
]

# ── Mapas de traducción ─────────────────────────────────────────

CATEGORY_LABELS: dict[str, str] = {
    "TOURIST_PASS":                "tour guiado",
    "MUSEUMS_AND_EXHIBITIONS":     "museo / exposición",
    "LUXURY_AND_SPECIAL_OCCASIONS":"ocasión especial / lujo",
    "CULINARY":                    "gastronomía",
    "CULTURAL_AND_THEME_TOURS":    "tour cultural",
    "SIGHTSEEING_ATTRACTION":      "atracción turística",
    "WALKING_TOUR":                "tour a pie",
    "NATURE":                      "naturaleza / exterior",
    "SHOWS_AND_MUSICALS":          "espectáculo en vivo",
    "ARTS_AND_CULTURE":            "arte y cultura",
}

ATTR_LABELS: dict[str, str] = {
    "FAMILY_FRIENDLY":  "apta para familia con niños",
    "ADULTS_ONLY":      "exclusiva para adultos (≥18)",
    "COUPLES":          "ideal para parejas",
    "ROMANTIC":         "romántica",
    "LUXURY":           "experiencia de lujo",
    "ECO_FRIENDLY":     "ecoamigable / al aire libre",
    "OUTDOOR":          "al aire libre",
    "GROUP_FRIENDLY":   "apta para grupos",
}

ACCESS_LABELS: dict[str, str] = {
    "WHEELCHAIR_ACCESSIBLE":               "accesible en silla de ruedas",
    "LIMITED_MOBILITY_ACCESSIBLE":         "accesible movilidad reducida",
    "STROLLER_OR_PRAM_ACCESSIBLE":         "accesible con coche de bebé",
    "PUBLIC_TRANSPORTATION_NEARBY":        "transporte público cercano",
}

# ── Helpers ─────────────────────────────────────────────────────

def clean_html(s: str) -> str:
    s = re.sub(r"<[^>]+>", " ", s or "")
    s = re.sub(r"&amp;", "&", s)
    s = re.sub(r"&#34;", '"', s)
    s = re.sub(r"&#43;", "+", s)
    s = re.sub(r"&nbsp;", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def split_items(raw: str) -> list[str]:
    """
    Splits space-separated capitalized phrases into individual items.
    e.g. "4 wine tastings Professional guide" -> ["4 wine tastings", "Professional guide"]
    Also handles existing bullet/newline separators.
    """
    if not raw:
        return []
    # If already has separators, use them
    if "\n" in raw or "•" in raw or "; " in raw:
        parts = re.split(r"[\n•;]+", raw)
        return [p.strip() for p in parts if p.strip()]
    # Split on capital letters that start a new word group (heuristic)
    # Insert separator before capital letters preceded by a lowercase letter or digit
    split = re.sub(r"(?<=[a-záéíóúñ\d])(?=[A-ZÁÉÍÓÚÑ])", "|||", raw)
    parts = [p.strip() for p in split.split("|||") if p.strip()]
    return parts if len(parts) > 1 else [raw]

def translate_list(raw: str, table: dict[str, str]) -> list[str]:
    """Split a comma/pipe-separated list and translate each token."""
    if not raw:
        return []
    tokens = [t.strip() for t in re.split(r"[,|]", raw) if t.strip()]
    return [table.get(t, t.lower().replace("_", " ")) for t in tokens]

def duration_label(h: float | None, m: float | None) -> str:
    h = int(h or 0)
    m = int(m or 0)
    parts = []
    if h:
        parts.append(f"{h} hora{'s' if h > 1 else ''}")
    if m:
        parts.append(f"{m} minutos")
    return " y ".join(parts) if parts else "variable"

def parse_min_age(val) -> int:
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return 0

def should_skip(title: str) -> bool:
    return any(kw in title for kw in SKIP_KEYWORDS)

# ── Perfiles de visitante (inferidos de atributos) ──────────────

def infer_profiles(attrs: list[str], min_age: int, cats: list[str]) -> list[str]:
    profiles = []
    has_family  = "apta para familia con niños" in attrs
    adults_only = "exclusiva para adultos (≥18)" in attrs or min_age >= 18
    couples     = "ideal para parejas" in attrs or "romántica" in attrs
    luxury      = "experiencia de lujo" in attrs or "ocasión especial / lujo" in cats
    groups      = "apta para grupos" in attrs
    gastro      = "gastronomía" in cats

    if adults_only:
        profiles.append("Adultos apasionados por el vino que buscan una cata profunda")
    if couples:
        profiles.append("Parejas que buscan una experiencia romántica o memorable")
    if luxury:
        profiles.append("Viajeros que buscan lo más exclusivo y premium")
    if has_family and not adults_only:
        profiles.append("Familias con niños (menores no participan en catas de vino)")
    if groups:
        profiles.append("Grupos de amigos o celebraciones corporativas")
    if gastro:
        profiles.append("Amantes de la gastronomía y el maridaje")
    if not profiles:
        profiles.append("Visitantes que se acercan al enoturismo por primera vez")
    return profiles

# ── Qué esperar (resumen ejecutivo del recorrido) ───────────────

WHAT_TO_EXPECT_OVERRIDES: dict[str, str] = {
    "Visita Guiada Standard": (
        "Recorrido de 70 minutos por el parque centenario, la Bodega Subterránea "
        "Casillero del Diablo y el Jardín de Variedades, con degustación de 4 vinos. "
        "Es el tour de entrada ideal: completo, asequible y bien explicado. "
        "No incluye almuerzo ni el museo sensorial completo."
    ),
    "Visita Guiada Premium": (
        "Recorrido de 2 horas por todos los espacios emblemáticos: parque, "
        "Museo Sensorial Casillero del Diablo, Bodega de Guarda El Alto, Mirador "
        "y Jardín de Variedades, con 4 copas de vinos premium. Ideal como experiencia "
        "completa para conocer la viña en profundidad."
    ),
    "Experiencia Nocturna Casillero del Diablo + Cena": (
        "Recorrido nocturno guiado por el enigmático Don Isidro, con visita al parque "
        "y a la Bodega Subterránea bajo las estrellas. Incluye cena de 3 tiempos en "
        "Bodega 1883 y degustación de vinos Casillero del Diablo. Experiencia íntima, "
        "máximo 25 personas. Prohibido menores de 18 años en la sala de cata."
    ),
    "Experiencia Vendimia Concha y Toro 2026": (
        "Vivencia estacional de la cosecha: pisada de uva, recorrido por el viñedo "
        "durante la vendimia, degustación y almuerzo. Solo disponible en temporada "
        "(febrero-marzo). Perfecta para quienes quieren vivir el proceso del vino desde dentro."
    ),
    "Tiny Wine Concerts": (
        "Concierto en vivo con formato íntimo en el viñedo, incluye vino y queso. "
        "Combinación de música en directo y enoturismo al atardecer. Ideal para "
        "quienes buscan una tarde diferente más allá del tour estándar."
    ),
}

# ── Lectora principal ───────────────────────────────────────────

def load_activities() -> list[dict]:
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb.active
    hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    def g(row: int, col: str) -> str:
        if col not in hdrs:
            return ""
        v = ws.cell(row, hdrs[col]).value
        return str(v).strip() if v else ""

    activities = []
    for r in range(2, ws.max_row + 1):
        title = g(r, "Title")
        if not title or should_skip(title):
            continue

        cats_raw  = g(r, "Categories")
        attrs_raw = g(r, "Attributes")
        access_raw = g(r, "KnowBeforeYouGo") + "," + g(r, "Inclusions")
        min_age   = parse_min_age(g(r, "Minimum age"))
        dur_h     = g(r, "Duration hours")
        dur_m     = g(r, "Duration minutes")

        cats    = translate_list(cats_raw, CATEGORY_LABELS)
        attrs   = translate_list(attrs_raw, ATTR_LABELS)
        access  = translate_list(access_raw, ACCESS_LABELS)
        access  = [a for a in access if a in ACCESS_LABELS.values()]  # filter noise

        desc     = clean_html(g(r, "Description"))
        included = clean_html(g(r, "Included"))
        excluded = clean_html(g(r, "Excluded"))
        attention= clean_html(g(r, "Attention"))
        excerpt  = clean_html(g(r, "Excerpt"))
        is_private = g(r, "Private experience").upper() == "TRUE"

        activities.append({
            "id":         g(r, "ID"),
            "title":      title,
            "duration":   duration_label(float(dur_h) if dur_h else None,
                                         float(dur_m) if dur_m else None),
            "categories": cats,
            "attributes": attrs,
            "access":     access,
            "min_age":    min_age,
            "desc":       desc,
            "included":   included,
            "excluded":   excluded,
            "attention":  attention,
            "excerpt":    excerpt,
            "is_private": is_private,
            "profiles":   infer_profiles(attrs, min_age, cats),
        })
    return activities


# ── Generador de Markdown ───────────────────────────────────────

def build_markdown(activities: list[dict]) -> str:
    lines: list[str] = []

    lines += [
        "# Guía de Experiencias — Centro del Vino Concha y Toro",
        "",
        "Este documento describe cada experiencia disponible desde el punto de vista "
        "del visitante: qué tipo de persona la disfrutará más, qué incluye, qué esperar "
        "y qué tener en cuenta. Úsalo para recomendar la experiencia más adecuada a cada perfil.",
        "",
        "---",
        "",
        "## Índice de Experiencias",
        "",
    ]
    for i, a in enumerate(activities, 1):
        priv = " *(privada)*" if a["is_private"] else ""
        lines.append(f"{i}. [{a['title']}](#{a['title'].lower().replace(' ','_').replace('+','').replace('/','').replace('(','').replace(')','').replace(',','').replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('ñ','n')}){priv}")
    lines += ["", "---", ""]

    for a in activities:
        # ── Encabezado ─────────────────────────────────────
        priv_tag = " — Privada" if a["is_private"] else ""
        lines += [f"## {a['title']}{priv_tag}", ""]

        # ── Snapshot ───────────────────────────────────────
        lines.append(f"**Duración:** {a['duration']}  ")
        if a["categories"]:
            lines.append(f"**Tipo:** {', '.join(dict.fromkeys(a['categories']))}  ")
        if a["min_age"] > 0:
            lines.append(f"**Edad mínima:** {a['min_age']} años  ")
        if a["access"]:
            lines.append(f"**Accesibilidad:** {', '.join(dict.fromkeys(a['access']))}  ")
        lines.append("")

        # ── Descripción ────────────────────────────────────
        # Use override summary if available (exact title match), else use description
        summary = WHAT_TO_EXPECT_OVERRIDES.get(a["title"])
        if not summary:
            # Try prefix match only if title starts with the key
            for key, text in WHAT_TO_EXPECT_OVERRIDES.items():
                if a["title"] == key or a["title"].startswith(key + " "):
                    summary = text
                    break
        if not summary and a["desc"]:
            # Take the first 600 chars of the clean description
            summary = a["desc"][:600].rsplit(" ", 1)[0] + "…" if len(a["desc"]) > 600 else a["desc"]

        if summary:
            lines += ["### Qué es esta experiencia", "", summary, ""]

        # ── Para quién es ──────────────────────────────────
        if a["profiles"]:
            lines += ["### Para quién es ideal", ""]
            for p in a["profiles"]:
                lines.append(f"- {p}")
            lines.append("")

        # ── Qué incluye / no incluye ───────────────────────
        if a["included"]:
            items = split_items(a["included"])
            lines += ["### Qué incluye", ""]
            for item in items:
                lines.append(f"- {item}")
            lines.append("")

        if a["excluded"]:
            items = split_items(a["excluded"])
            lines += ["### Qué NO incluye", ""]
            for item in items:
                lines.append(f"- {item}")
            lines.append("")

        # ── Importante / restricciones ─────────────────────
        notices = []
        if a["min_age"] >= 18:
            notices.append("Exclusiva para mayores de 18 años. Sala de cata cerrada para menores.")
        elif a["min_age"] > 0:
            notices.append(f"Menores de {a['min_age']} años no admitidos en algunas secciones.")
        if a["attention"]:
            att_items = split_items(a["attention"])
            for item in att_items[:5]:  # limit to 5 items
                if len(item) > 10:  # skip very short fragments
                    notices.append(item)

        if notices:
            lines += ["### Importante antes de reservar", ""]
            for n in notices:
                lines.append(f"- {n}")
            lines.append("")

        lines += ["---", ""]

    return "\n".join(lines)


# ── Comparativa por perfil ──────────────────────────────────────

PROFILE_GUIDE = """
---

## Comparativa rápida por perfil de visitante

| Perfil | Experiencias recomendadas |
|--------|--------------------------|
| **Primera visita, quiero ver todo en poco tiempo** | Visita Guiada Standard (70 min) |
| **Primera visita, quiero profundidad** | Visita Guiada Premium (2h) |
| **Pareja, algo romántico o especial** | Experiencia Nocturna + Cena, Tasting Cellar Collection, Maridaje Terrunyo |
| **Amante del vino premium / coleccionista** | Cellar Collection, Casa Don Melchor, The New Wines, Marqués de Casa Concha |
| **Familia con niños** | Visita Guiada Standard o Premium (niños entran gratis, no catarán) |
| **Grupo de amigos / celebración** | Nocturna + Cena, Premium + Maridaje La Gran Barra, Tiny Wine Concerts |
| **Gastronomía + vino** | Maridaje Terrunyo, Nocturna + Cena, Premium + Almuerzo Bodega 1883 |
| **Evento especial / propuesta / aniversario** | Nocturna + Cena, Privada, Casa Don Melchor |
| **Temporada de vendimia (feb–mar)** | Experiencia Vendimia 2026 |
| **Tarde cultural con música** | Tiny Wine Concerts |

---

## Preguntas frecuentes de asesoría

**¿Cuál es la diferencia entre Standard y Premium?**
La Visita Standard (70 min) es más corta, recorre el parque y la Bodega Casillero del Diablo con 4 catas. 
La Premium (2h) añade el Museo Sensorial, la Bodega de Guarda El Alto y el Mirador, con más profundidad cultural y las mismas 4 catas.

**¿Pueden venir niños?**
Sí en la mayoría de las experiencias. Los menores de 18 años no participan en las catas de vino pero sí en el recorrido. 
En la Experiencia Nocturna + Cena y otras con sala de cata cerrada, no se admiten menores de 18.

**¿Qué experiencia es mejor para alguien que no sabe nada de vino?**
La Visita Guiada Standard o Premium. Los guías están preparados para todos los niveles, desde principiantes hasta expertos.

**¿Hay algo para quien ya ha venido antes?**
Sí: las experiencias premium de degustación (Cellar Collection, Casa Don Melchor, Terrunyo, The New Wines) están pensadas para quienes ya conocen la viña y quieren ir más allá.

**¿Se puede venir sin tour y solo comer?**
Sí. El restaurante Bodega 1883 y La Gran Barra aceptan reservas independientes del tour.

**¿Las experiencias privadas son más caras?**
Sí, pero permiten horarios flexibles, grupos reducidos y una atención personalizada. Ideales para celebraciones o eventos corporativos.
"""

# ── Main ────────────────────────────────────────────────────────

def main() -> None:
    print(f"Leyendo: {XLSX}")
    activities = load_activities()
    print(f"Experiencias procesadas: {len(activities)}")

    md = build_markdown(activities)
    md += PROFILE_GUIDE

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(md, encoding="utf-8")
    print(f"Generado: {OUT}")
    print(f"Tamaño: {len(md):,} caracteres / ~{len(md)//4:,} tokens estimados")


if __name__ == "__main__":
    main()
